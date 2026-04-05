from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Color definitions
DARK_BLUE = PatternFill('solid', fgColor='1F4E79')
WHITE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
LIGHT_BLUE = PatternFill('solid', fgColor='D6E4F0')
LIGHT_ORANGE = PatternFill('solid', fgColor='FCE4D6')
LIGHT_GRAY = PatternFill('solid', fgColor='E2EFDA')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_headers(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = DARK_BLUE
        cell.font = WHITE_FONT
        cell.alignment = HEADER_ALIGN

def style_body(ws, num_rows, num_cols, score_col=None):
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if score_col:
            score_cell = ws.cell(row=row, column=score_col)
            try:
                val = float(score_cell.value)
                fill = GREEN_FILL if val >= 7.0 else YELLOW_FILL if val >= 5.0 else RED_FILL
                for col in range(1, num_cols + 1):
                    ws.cell(row=row, column=col).fill = fill
            except (ValueError, TypeError):
                pass
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col).border = THIN_BORDER

def auto_width(ws, num_cols):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 35)

# ============ SHEET 1: Master Ranking ============
ws1 = wb.active
ws1.title = "Master Ranking"
ws1.sheet_properties.tabColor = "00B050"

headers1 = ['Rank', 'Stock', 'Watchlist', 'Setup Type', 'Direction', '4H Score', 'Confidence', 'Entry Price', 'Stop Loss', 'Target', 'Key Edge']
ws1.append(headers1)

master_data = [
    [1, 'NSE:VEDL', 'Technical', 'BO-1', 'BUY', 8.0, 'HIGH', '677 (dip to Filter)', '644 (4H Tunnel)', 'Trail', 'All 3 TFs aligned + Wave > Tunnel on 4H'],
    [2, 'NSE:ONGC', 'Nifty50', 'BO-3', 'BUY', 7.5, 'HIGH', '286 (Filter) / 283 (Wave)', '281 (below Wave)', 'Trail', 'Perfect 4H stacking, healthy pullback to Filter'],
    [3, 'NSE:ANGELONE', 'Technical', 'BO-1', 'BUY', 7.0, 'MED-HIGH', '241 or dip to 238', '231 (4H Wave)', 'Trail', '4H at Tunnel breakout, AO already positive'],
    [4, 'NSE:HINDALCO', 'Nifty50', 'BO-1', 'BUY', 7.0, 'MED-HIGH', '906 (dip to Tunnel)', '895 (Wave)', 'Trail', 'Fresh 4H Tunnel breakout, strong AO +19'],
    [5, 'NSE:CUMMINSIND', 'Technical', 'BO-2', 'BUY', 6.5, 'MED-HIGH', '4617 (dip to Filter)', '4477 (4H Tunnel)', 'Trail', '4H Wave > Tunnel, re-entry setup'],
    [6, 'NSE:ADANIPORTS', 'Nifty50', 'BO-1 dev', 'BUY', 6.5, 'MED-HIGH', 'Above 1391 (Tunnel)', '1359 (Wave)', 'Trail', 'AO positive, just below Tunnel'],
    [7, 'NSE:CGPOWER', 'Nifty50', 'BO-1 comp', 'BUY', 6.0, 'MED-HIGH', 'Above 686 (Tunnel)', '670 (Wave)', 'Trail', 'Coiled spring, AO about to cross zero'],
    [8, 'NSE:GAIL', 'Nifty50', 'BO-1 watch', 'BUY', 5.5, 'MEDIUM', 'Above 146 (Tunnel)', '139', 'Trail', 'AO positive, near Tunnel'],
    [9, 'NSE:TATAPOWER', 'Nifty50', 'BO-1 Tunnel', 'BUY', 5.5, 'MEDIUM', 'Above 386 (Tunnel)', '380', 'Trail', 'Tight compression at Tunnel'],
    [10, 'NSE:AMBUJACEM', 'Nifty50', 'Bounce', 'BUY', 5.5, 'MEDIUM', 'Above 415', '411', '438 (Tunnel)', 'AO just crossed positive'],
    [11, 'NSE:ADANIGREEN', 'Nifty50', 'BO-1 watch', 'BUY', 5.5, 'MEDIUM', 'Above 863 (Tunnel)', '838', 'Trail', 'AO near zero rising, near Tunnel'],
    [12, 'NSE:KALYANKJIL', 'Technical', 'BO-1', 'BUY', 5.0, 'MEDIUM', '389 (15m Filter)', '384 (Tunnel)', 'Trail', '15m/1H confirmed but 4H opposes - scalp'],
    [13, 'NSE:SUZLON', 'Nifty50', 'BO-1 watch', 'BUY', 5.0, 'MEDIUM', 'Above 42 (Tunnel)', '40', 'Trail', 'Extreme compression, tiny breakout distance'],
    [14, 'NSE:INDIGO', 'Nifty50', 'Bounce', 'BUY', 5.0, 'MEDIUM', 'Current ~4170', '4100', '4281 (Tunnel)', 'AO improving fast'],
    [15, 'NSE:NYKAA', 'Technical', 'BO-1 (15m)', 'BUY', 4.5, 'MEDIUM', '242 (15m Filter)', '239', '251 (4H Tunnel)', '15m confirmed, 4H opposes - scalp'],
    [16, 'NSE:INFY', 'Technical', 'BO-3 (15m)', 'BUY', 4.0, 'LOW-MED', '1289 (15m Filter)', '1277', '1406 (4H Tunnel)', '15m strong but 4H deeply bearish'],
]
for row in master_data:
    ws1.append(row)
style_headers(ws1, len(headers1))
style_body(ws1, len(master_data) + 1, len(headers1), score_col=6)
auto_width(ws1, len(headers1))

# ============ SHEET 2: 4H Technical ============
ws2 = wb.create_sheet("4H Technical")
ws2.sheet_properties.tabColor = "4472C4"

headers2 = ['Stock', 'Price', 'Filter(12)', 'Wave(34)', 'Tunnel(144)', 'Tunnel(169)', 'AO', 'AO Signal', 'Bungee(76)', 'Bungee(34)', 'Bungee(16)', 'Bungee(8)', 'Williams%R', 'RSI(2)', '4H Trend']
ws2.append(headers2)

tech_data = [
    ['NSE:VEDL', 686.50, 668.44, 671.09, 643.88, 633.34, -2.58, -10.81, 53.2, 51.2, 83.7, 74.2, 92.1, 78.3, 'ABOVE TUNNEL - BO-1'],
    ['NSE:ANGELONE', 241.10, 235.34, 231.38, 240.11, 241.21, 11.40, 11.22, 46.7, 87.4, 82.7, 74.4, 81.8, 68.8, 'AT TUNNEL - BO-1'],
    ['NSE:CUMMINSIND', 4650.40, 4615.80, 4630.52, 4477.41, 4439.61, -48.18, -51.55, 63.4, 43.4, 59.6, 55.6, 90.8, 71.5, 'ABOVE TUNNEL - BO-2'],
    ['NSE:PERSISTENT', 5222.40, 4987.53, 4893.99, 5358.18, 5415.66, 261.43, 196.06, 42.3, 94.2, 93.2, 88.6, 89.9, 91.4, 'BELOW TUNNEL - Recovery'],
    ['NSE:KALYANKJIL', 392.90, 385.40, 386.52, 417.57, 424.27, 0.91, 0.76, 33.1, 61.6, 74.4, 63.2, 86.5, 77.0, 'BELOW TUNNEL - Bounce'],
    ['NSE:OBEROIRLTY', 1512.00, 1464.48, 1464.85, 1532.20, 1544.35, -1.18, -3.29, 44.6, 77.0, 81.7, 79.6, 98.9, 94.6, 'BELOW TUNNEL - Recovery'],
    ['NSE:INFY', 1298.00, 1274.15, 1279.78, 1406.45, 1421.87, 9.53, 3.13, 18.7, 63.2, 76.2, 65.8, 88.1, 85.1, 'DEEP DOWNTREND'],
    ['NSE:NYKAA', 245.82, 240.14, 243.42, 251.19, 250.92, -3.28, -4.73, 16.5, 32.5, 57.4, 60.3, 90.4, 91.2, 'BELOW TUNNEL - Bounce'],
    ['NSE:M&M', 3007.10, 3027.16, 3101.29, 3355.99, 3376.55, -73.73, -78.21, 13.7, 23.6, 31.6, 34.2, 55.5, 53.9, 'DEEP DOWNTREND'],
    ['NSE:AXISBANK', 1198.10, 1195.30, 1227.31, 1272.84, 1270.16, -39.61, -43.23, 16.3, 20.7, 42.3, 50.1, 82.8, 64.6, 'DEEP DOWNTREND'],
    ['NSE:ICICIBANK', 1216.50, 1227.39, 1262.08, 1334.24, 1340.29, -45.65, -42.29, 7.7, 12.4, 19.5, 22.7, 62.3, 63.7, 'DEEP DOWNTREND'],
    ['NSE:BANKBARODA', 250.00, 257.51, 271.87, 286.14, 285.57, -25.04, -23.47, 8.4, 11.3, 15.3, 20.7, 59.0, 55.2, 'DEEP DOWNTREND'],
    ['NSE:CANBK', 127.00, 128.95, 134.53, 142.28, 141.90, -8.67, -8.14, 10.8, 16.2, 22.4, 28.0, 67.6, 60.5, 'DEEP DOWNTREND'],
    ['NSE:LTF', 242.00, 247.59, 257.86, 276.35, 276.23, -15.90, -14.24, 8.4, 12.6, 16.7, 19.1, 44.7, 53.3, 'DEEP DOWNTREND'],
    ['NSE:PNBHOUSING', 779.65, 781.08, 787.15, 837.92, 845.03, -4.94, -0.29, 36.3, 56.5, 52.7, 47.0, 65.6, 57.4, 'BELOW TUNNEL'],
    ['NSE:SWIGGY', 275.15, 271.44, 281.45, 321.94, 328.53, -15.61, -16.98, 11.4, 25.3, 35.3, 53.9, 90.9, 93.9, 'DEEP DOWNTREND'],
    ['NSE:LT', 3610.00, 3569.70, 3639.00, 3866.61, 3873.55, -14.43, -44.01, 28.0, 43.4, 76.3, 55.0, 65.8, 61.4, 'BELOW TUNNEL'],
]
for row in tech_data:
    ws2.append(row)
style_headers(ws2, len(headers2))
style_body(ws2, len(tech_data) + 1, len(headers2))
auto_width(ws2, len(headers2))

# ============ SHEET 3: 4H Nifty50 ============
ws3 = wb.create_sheet("4H Nifty50")
ws3.sheet_properties.tabColor = "ED7D31"

ws3.append(headers2)

nifty_data = [
    ['NSE:ONGC', 287.20, 286.46, 282.69, 274.01, 273.00, 4.02, 4.89, 81.4, 77.5, 47.5, 67.3, 59.5, 33.8, 'ABOVE TUNNEL - BO-3'],
    ['NSE:HINDALCO', 916.25, 906.75, 895.08, 905.86, 908.41, 19.30, 14.49, 75.6, 93.5, 88.9, 94.0, 97.1, 89.2, 'BO-1 BREAKOUT'],
    ['NSE:ADANIPORTS', 1377.60, 1366.16, 1358.83, 1390.51, 1398.32, 5.71, 1.55, 65.7, 70.1, 70.1, 73.2, 90.2, 93.9, 'NEAR TUNNEL - BO-1 dev'],
    ['NSE:CGPOWER', 680.05, 674.35, 673.97, 685.66, 685.70, -1.36, -5.01, 46.1, 67.8, 74.3, 82.8, 95.1, 94.5, 'COMPRESSION near Tunnel'],
    ['NSE:GAIL', 141.73, 140.33, 139.89, 146.26, 147.57, 1.14, 0.38, 39.0, 83.5, 76.1, 90.4, 95.5, 95.1, 'BELOW TUNNEL - AO +ve'],
    ['NSE:TATAPOWER', 385.00, 379.76, 382.99, 385.74, 385.04, -5.71, -8.86, 25.1, 54.4, 64.2, 82.8, 78.2, 65.6, 'COMPRESSION at Tunnel'],
    ['NSE:AMBUJACEM', 418.45, 414.83, 414.04, 437.92, 443.85, 0.40, -0.76, 44.5, 65.8, 78.2, 68.3, 75.6, 51.2, 'BELOW TUNNEL - AO +ve'],
    ['NSE:ADANIGREEN', 856.00, 840.28, 838.84, 863.29, 868.97, -1.63, -6.26, 40.8, 73.3, 81.0, 75.4, 78.2, 69.7, 'NEAR TUNNEL - AO ~0'],
    ['NSE:SUZLON', 40.94, 40.61, 40.78, 41.61, 41.89, -0.59, -0.74, 40.3, 44.8, 67.5, 65.5, 96.7, 94.7, 'COMPRESSION at Tunnel'],
    ['NSE:INDIGO', 4170.00, 4132.34, 4135.04, 4281.20, 4319.50, -33.32, -72.77, 48.7, 60.6, 60.6, 78.2, 81.3, 62.1, 'BELOW TUNNEL - Recovery'],
    ['NSE:DLF', 522.25, 513.03, 516.77, 547.60, 553.95, -5.00, -12.95, 34.4, 59.6, 88.3, 94.7, 89.4, 71.9, 'BELOW TUNNEL - Bounce'],
    ['NSE:GODREJPROP', 1508.30, 1491.38, 1504.51, 1584.97, 1599.97, -24.75, -39.49, 25.7, 42.9, 67.0, 83.9, 89.6, 74.9, 'BELOW TUNNEL'],
    ['NSE:INDHOTEL', 582.75, 580.80, 587.81, 614.33, 619.10, -13.09, -16.92, 20.5, 29.7, 58.4, 76.6, 96.4, 93.1, 'BELOW TUNNEL - Bounce'],
    ['NSE:BHARTIARTL', 1789.70, 1781.60, 1796.17, 1839.56, 1849.38, -34.70, -41.44, 26.4, 37.1, 58.0, 78.0, 85.9, 70.3, 'BELOW TUNNEL'],
    ['NSE:INDUSTOWER', 424.85, 421.93, 423.43, 433.08, 434.52, -3.91, -6.51, 37.3, 53.6, 53.6, 82.3, 97.4, 87.6, 'NEAR TUNNEL'],
    ['NSE:POWERGRID', 289.95, 290.12, 293.14, 295.81, 295.32, -6.53, -7.44, 24.1, 29.0, 29.0, 59.9, 87.6, 92.8, 'NEAR TUNNEL'],
    ['NSE:RELIANCE', 1350.50, 1353.43, 1365.93, 1390.54, 1393.67, -20.97, -25.68, 21.2, 24.1, 43.8, 55.1, 75.7, 48.0, 'NEAR TUNNEL'],
    ['NSE:BPCL', 278.15, 277.68, 280.59, 305.75, 310.61, -6.07, -8.20, 25.8, 36.1, 51.4, 67.8, 85.2, 91.5, 'BELOW TUNNEL'],
    ['NSE:APOLLOHOSP', 7299.00, 7310.10, 7367.19, 7460.11, 7462.88, -143.96, -186.33, 42.1, 36.6, 39.5, 81.3, 67.1, 39.7, 'BELOW TUNNEL'],
    ['NSE:MOTHERSON', 106.81, 106.58, 107.93, 114.17, 115.06, -2.76, -3.62, 23.8, 33.4, 54.2, 73.1, 93.8, 86.1, 'BELOW TUNNEL'],
    ['NSE:SHREECEM', 23115.00, 23119.79, 23277.32, 23951.33, 24124.14, -425.87, -518.40, 32.2, 29.1, 58.5, 70.3, 74.6, 46.2, 'BELOW TUNNEL'],
    ['NSE:MAXHEALTH', 946.50, 949.32, 958.43, 989.38, 994.41, -20.44, -22.68, 26.2, 28.2, 34.6, 53.2, 79.9, 59.7, 'BELOW TUNNEL'],
    ['NSE:ULTRACEMCO', 10600.00, 10628.06, 10754.49, 11229.35, 11327.80, -327.21, -365.79, 26.6, 28.4, 33.1, 60.9, 75.0, 52.9, 'BELOW TUNNEL'],
    ['NSE:HINDPETRO', 326.00, 328.16, 333.22, 360.19, 366.26, -13.50, -13.92, 20.8, 19.3, 27.9, 36.2, 92.4, 75.9, 'DEEP DOWNTREND'],
    ['NSE:IOC', 134.13, 134.07, 136.33, 148.19, 150.04, -3.86, -4.59, 17.4, 26.5, 38.6, 58.9, 89.4, 76.5, 'DEEP DOWNTREND'],
    ['NSE:BHARATFORG', 1651.40, 1655.44, 1679.16, 1724.02, 1723.59, -52.12, -56.91, 14.7, 23.9, 28.1, 44.8, 84.7, 83.8, 'DEEP DOWNTREND'],
    ['NSE:ASHOKLEY', 148.44, 149.13, 155.50, 172.92, 175.27, -12.13, -13.70, 10.7, 17.2, 26.0, 54.8, 87.7, 77.5, 'DEEP DOWNTREND'],
    ['NSE:GRASIM', 2564.10, 2560.52, 2579.36, 2653.22, 2667.03, -45.83, -51.51, 23.1, 30.8, 48.0, 55.3, 66.3, 38.7, 'DOWNTREND'],
    ['NSE:NTPC', 359.65, 361.12, 367.64, 373.75, 373.35, -13.65, -14.86, 19.3, 24.6, 24.6, 49.5, 74.2, 50.6, 'DOWNTREND'],
]
for row in nifty_data:
    ws3.append(row)
style_headers(ws3, len(headers2))
style_body(ws3, len(nifty_data) + 1, len(headers2))
auto_width(ws3, len(headers2))

# ============ SHEET 4: Setup Legend ============
ws4 = wb.create_sheet("Setup Legend")
ws4.sheet_properties.tabColor = "A5A5A5"

headers4 = ['Setup Type', 'Category', 'Description', 'Entry Rule', 'Stop Loss Rule', 'Target Rule']
ws4.append(headers4)

legend_data = [
    ['PW', 'End-of-Trend', 'Price extended from Wave/Tunnel with AO divergence', 'Reversal candle or Filter cross', 'Beyond extreme high/low', 'Wave (34 EMA)'],
    ['FG', 'End-of-Trend', 'Fill the gap between Wave and Tunnel after PW', 'Reversal candle + Filter close after Wave 2/B', 'Beyond Wave 2/B extreme', 'Tunnel'],
    ['BO-1', 'Trend Following', 'Wave crosses Tunnel - first breakout', 'Close above/below Filter when Wave crosses Tunnel', 'Beyond 34 EMA', 'Trail with Three Profiteers'],
    ['BO-2', 'Trend Following', 'Wave bounces off Tunnel - re-entry after correction', 'Cross Wave + close above Filter + steep Filter angle', 'Beyond 34 EMA', 'Trail with Three Profiteers'],
    ['BO-3', 'Trend Following', 'Price bounces off Wave - pullback buy/sell', 'Bounce off Wave with reversal candle or Filter close', 'Below/above Wave', 'Trail with Three Profiteers'],
    ['BO-4', 'Trend Following', 'Price bounces off Tunnel after steep Wave 3', 'Bounce off Tunnel with reversal candle', 'Below/above Tunnel', 'Trail with Three Profiteers'],
]
for row in legend_data:
    ws4.append(row)
style_headers(ws4, len(headers4))
style_body(ws4, len(legend_data) + 1, len(headers4))
auto_width(ws4, len(headers4))

# ============ SHEET 5: Confidence Scale ============
ws5 = wb.create_sheet("Confidence Scale")
ws5.sheet_properties.tabColor = "7030A0"

headers5 = ['Score Range', 'Rating', 'Action', 'Position Size']
ws5.append(headers5)

scale_data = [
    ['8 - 10', 'HIGH', 'Execute with full size', '3 positions (Three Profiteers)'],
    ['6 - 7.9', 'MEDIUM-HIGH', 'Execute with 2 positions; add 3rd on confirmation', '2 positions initially'],
    ['4 - 5.9', 'MEDIUM', 'Execute with 1 position; wait for confirmation', '1 position'],
    ['2 - 3.9', 'LOW', 'Monitor only - do not trade', 'Place alerts only'],
    ['0 - 1.9', 'PASS', 'No valid setup', 'Ignore'],
]
for row in scale_data:
    ws5.append(row)
style_headers(ws5, len(headers5))

fills = [GREEN_FILL, LIGHT_BLUE, YELLOW_FILL, LIGHT_ORANGE, RED_FILL]
for i, row_data in enumerate(scale_data):
    for col in range(1, len(headers5) + 1):
        cell = ws5.cell(row=i + 2, column=col)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = fills[i]

auto_width(ws5, len(headers5))

# Freeze panes on data sheets
ws1.freeze_panes = 'A2'
ws2.freeze_panes = 'B2'
ws3.freeze_panes = 'B2'

output_path = r'C:\Users\91700\tradingview-mcp-jackson\Wavy_Tunnel_Analysis_2026-04-04.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
