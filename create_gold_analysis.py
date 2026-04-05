from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DARK_BLUE = PatternFill('solid', fgColor='1F4E79')
DARK_GOLD = PatternFill('solid', fgColor='BF8F00')
WHITE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
LIGHT_BLUE = PatternFill('solid', fgColor='D6E4F0')
LIGHT_ORANGE = PatternFill('solid', fgColor='FCE4D6')
LIGHT_GRAY = PatternFill('solid', fgColor='E2EFDA')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_headers(ws, num_cols, fill=DARK_BLUE):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.alignment = HEADER_ALIGN

def style_body(ws, num_rows, num_cols):
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

def auto_width(ws, num_cols):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 25)

# ============================================================
# DATA: All 18 symbols across 3 timeframes
# ============================================================
# Indicator mapping:
#   Filter = Fast EMA(12)
#   Wave = Tunnel High(34) / Tunnel Low(34) / Tunnel Close(34)
#   Tunnel = Wave EMA(144) / Wave EMA(169)
#   AO = Awesome Oscillator
#   Bungee = 76/34/16/8 stochastics + Williams %R + RSI(2)

data = [
    {
        "symbol": "OANDA:XAUUSD", "name": "Gold Spot", "last": 4676.745,
        "tf15": {"filter": 4667.664, "wave_h": 4665.406, "wave_l": 4651.111, "wave_c": 4658.357, "tunnel_h": 4667.921, "tunnel_l": 4665.937, "ao": 17.202, "ao_sig": 19.868, "b76": 71.7, "b34": 75.8, "b16": 69.8, "b8": 54.3, "wr": 91.8, "rsi2": 84.2},
        "tf1h": {"filter": 4659.757, "wave_h": 4685.494, "wave_l": 4649.796, "wave_c": 4668.425, "tunnel_h": 4615.545, "tunnel_l": 4616.786, "ao": -31.330, "ao_sig": -41.016, "b76": 57.8, "b34": 48.1, "b16": 81.3, "b8": 77.1, "wr": 82.0, "rsi2": 85.1},
        "tf4h": {"filter": 4667.395, "wave_h": 4654.280, "wave_l": 4580.192, "wave_c": 4619.248, "tunnel_h": 4763.816, "tunnel_l": 4785.230, "ao": 81.700, "ao_sig": 137.921, "b76": 57.2, "b34": 65.5, "b16": 46.5, "b8": 37.3, "wr": 49.8, "rsi2": 71.8},
    },
    {
        "symbol": "MCX:GOLD1!", "name": "Gold Futures MCX", "last": 149650,
        "tf15": {"filter": 149780, "wave_h": 149835, "wave_l": 149320, "wave_c": 149586, "tunnel_h": 149669, "tunnel_l": 149441, "ao": 644, "ao_sig": 756, "b76": 36.3, "b34": 66.1, "b16": 52.7, "b8": 36.8, "wr": 16.9, "rsi2": 4.0},
        "tf1h": {"filter": 149690, "wave_h": 150361, "wave_l": 149293, "wave_c": 149799, "tunnel_h": 148321, "tunnel_l": 148663, "ao": -1129, "ao_sig": -1629, "b76": 70.2, "b34": 35.8, "b16": 36.8, "b8": 65.9, "wr": 62.8, "rsi2": 36.9},
        "tf4h": {"filter": 149651, "wave_h": 149381, "wave_l": 147055, "wave_c": 148317, "tunnel_h": 151718, "tunnel_l": 151461, "ao": 5443, "ao_sig": 6775, "b76": 58.1, "b34": 79.9, "b16": 63.2, "b8": 32.4, "wr": 34.5, "rsi2": 41.2},
    },
    {
        "symbol": "TVC:DXY", "name": "US Dollar Index", "last": 100.185,
        "tf15": {"filter": 100.193, "wave_h": 100.167, "wave_l": 100.139, "wave_c": 100.154, "tunnel_h": 100.044, "tunnel_l": 100.032, "ao": 0.036, "ao_sig": 0.059, "b76": 85.6, "b34": 85.3, "b16": 67.3, "b8": 67.3, "wr": 65.5, "rsi2": 33.1},
        "tf1h": {"filter": 100.138, "wave_h": 100.083, "wave_l": 100.011, "wave_c": 100.053, "tunnel_h": 99.947, "tunnel_l": 99.927, "ao": 0.135, "ao_sig": 0.106, "b76": 93.5, "b34": 86.5, "b16": 94.7, "b8": 92.7, "wr": 65.5, "rsi2": 28.8},
        "tf4h": {"filter": 100.052, "wave_h": 100.068, "wave_l": 99.857, "wave_c": 99.973, "tunnel_h": 99.482, "tunnel_l": 99.372, "ao": 0.024, "ao_sig": -0.012, "b76": 71.2, "b34": 66.2, "b16": 92.3, "b8": 84.8, "wr": 86.3, "rsi2": 66.1},
    },
    {
        "symbol": "NSE:RELIANCE", "name": "Reliance Industries", "last": 1350.9,
        "tf15": {"filter": 1348.79, "wave_h": 1353.33, "wave_l": 1347.22, "wave_c": 1350.41, "tunnel_h": 1368.77, "tunnel_l": 1371.80, "ao": 1.41, "ao_sig": -1.94, "b76": 40.3, "b34": 48.7, "b16": 72.1, "b8": 67.2, "wr": 48.2, "rsi2": 42.3},
        "tf1h": {"filter": 1353.43, "wave_h": 1371.37, "wave_l": 1359.97, "wave_c": 1365.93, "tunnel_h": 1390.54, "tunnel_l": 1393.67, "ao": -20.97, "ao_sig": -25.68, "b76": 21.2, "b34": 24.1, "b16": 43.8, "b8": 55.1, "wr": 75.7, "rsi2": 48.0},
        "tf4h": {"filter": 1369.18, "wave_h": 1397.74, "wave_l": 1376.45, "wave_c": 1386.34, "tunnel_h": 1425.28, "tunnel_l": 1429.31, "ao": -35.50, "ao_sig": -34.92, "b76": 31.7, "b34": 23.2, "b16": 24.1, "b8": 26.8, "wr": 40.6, "rsi2": 44.4},
    },
    {
        "symbol": "MCX:SILVER1!", "name": "Silver Futures MCX", "last": 232600,
        "tf15": {"filter": 232162, "wave_h": 231784, "wave_l": 230521, "wave_c": 231172, "tunnel_h": 233090, "tunnel_l": 233070, "ao": 2913, "ao_sig": 2842, "b76": 42.6, "b34": 80.1, "b16": 70.7, "b8": 59.6, "wr": 34.2, "rsi2": 27.6},
        "tf1h": {"filter": 231492, "wave_h": 234561, "wave_l": 231851, "wave_c": 233138, "tunnel_h": 234979, "tunnel_l": 236253, "ao": -4385, "ao_sig": -6571, "b76": 53.3, "b34": 40.5, "b16": 40.5, "b8": 77.4, "wr": 79.4, "rsi2": 81.0},
        "tf4h": {"filter": 233909, "wave_h": 237564, "wave_l": 231416, "wave_c": 234768, "tunnel_h": 249066, "tunnel_l": 249543, "ao": 4259, "ao_sig": 8053, "b76": 38.6, "b34": 70.7, "b16": 43.0, "b8": 33.4, "wr": 41.6, "rsi2": 56.7},
    },
    {
        "symbol": "MCX:MCXCOPRDEX", "name": "MCX Copper Index", "last": 18047.88,
        "tf15": {"filter": 18050.80, "wave_h": 18047.50, "wave_l": 18018.88, "wave_c": 18032.88, "tunnel_h": 18062.09, "tunnel_l": 18059.50, "ao": 41.52, "ao_sig": 52.60, "b76": 37.8, "b34": 53.4, "b16": 37.8, "b8": 21.0, "wr": 40.8, "rsi2": 54.2},
        "tf1h": {"filter": 18036.64, "wave_h": 18096.61, "wave_l": 18032.98, "wave_c": 18065.30, "tunnel_h": 18073.05, "tunnel_l": 18107.51, "ao": -58.85, "ao_sig": -104.05, "b76": 52.0, "b34": 36.6, "b16": 38.0, "b8": 53.9, "wr": 54.0, "rsi2": 51.1},
        "tf4h": {"filter": 18077.77, "wave_h": 18142.53, "wave_l": 17980.50, "wave_c": 18069.82, "tunnel_h": 18755.83, "tunnel_l": 18855.19, "ao": 165.46, "ao_sig": 261.94, "b76": 54.2, "b34": 80.6, "b16": 34.5, "b8": 30.9, "wr": 38.1, "rsi2": 48.2},
    },
    {
        "symbol": "MCX:MCXENRGDEX", "name": "MCX Energy Index", "last": 9046.81,
        "tf15": {"filter": 9010.38, "wave_h": 8973.36, "wave_l": 8881.92, "wave_c": 8933.62, "tunnel_h": 8600.76, "tunnel_l": 8556.08, "ao": 54.81, "ao_sig": 79.76, "b76": 79.8, "b34": 62.6, "b16": 74.6, "b8": 59.6, "wr": 79.6, "rsi2": 83.1},
        "tf1h": {"filter": 8895.42, "wave_h": 8696.66, "wave_l": 8544.54, "wave_c": 8639.29, "tunnel_h": 8157.22, "tunnel_l": 8077.96, "ao": 470.71, "ao_sig": 485.91, "b76": 87.2, "b34": 83.9, "b16": 82.0, "b8": 64.2, "wr": 65.7, "rsi2": 64.2},
        "tf4h": {"filter": 8561.68, "wave_h": 8334.45, "wave_l": 8040.30, "wave_c": 8206.03, "tunnel_h": 7085.44, "tunnel_l": 6913.07, "ao": 593.66, "ao_sig": 380.01, "b76": 88.2, "b34": 91.4, "b16": 86.3, "b8": 86.2, "wr": 82.7, "rsi2": 96.4},
    },
    {
        "symbol": "MCX:MCXZINCDEX", "name": "MCX Zinc Index", "last": 28773.71,
        "tf15": {"filter": 28810.11, "wave_h": 28804.37, "wave_l": 28752.78, "wave_c": 28774.61, "tunnel_h": 28584.58, "tunnel_l": 28539.75, "ao": 39.65, "ao_sig": 51.29, "b76": 56.5, "b34": 52.2, "b16": 16.1, "b8": 14.9, "wr": 0.0, "rsi2": 17.5},
        "tf1h": {"filter": 28748.50, "wave_h": 28683.42, "wave_l": 28559.59, "wave_c": 28615.65, "tunnel_h": 28211.13, "tunnel_l": 28202.97, "ao": 212.56, "ao_sig": 184.74, "b76": 83.7, "b34": 67.0, "b16": 54.4, "b8": 49.7, "wr": 36.0, "rsi2": 34.4},
        "tf4h": {"filter": 28578.18, "wave_h": 28365.94, "wave_l": 28114.63, "wave_c": 28251.92, "tunnel_h": 28341.55, "tunnel_l": 28329.21, "ao": 795.13, "ao_sig": 802.65, "b76": 84.7, "b34": 87.0, "b16": 79.1, "b8": 66.2, "wr": 53.1, "rsi2": 49.7},
    },
    {
        "symbol": "MCX:MCXNGASDEX", "name": "MCX Natural Gas Index", "last": 851.44,
        "tf15": {"filter": 854.29, "wave_h": 861.55, "wave_l": 856.96, "wave_c": 858.77, "tunnel_h": 872.62, "tunnel_l": 874.98, "ao": -6.74, "ao_sig": -6.16, "b76": 14.6, "b34": 15.1, "b16": 36.0, "b8": 12.5, "wr": 4.1, "rsi2": 19.1},
        "tf1h": {"filter": 859.53, "wave_h": 875.55, "wave_l": 866.62, "wave_c": 870.50, "tunnel_h": 895.67, "tunnel_l": 898.44, "ao": -17.29, "ao_sig": -14.64, "b76": 6.0, "b34": 9.1, "b16": 16.8, "b8": 19.7, "wr": 14.7, "rsi2": 23.5},
        "tf4h": {"filter": 871.81, "wave_h": 902.61, "wave_l": 883.59, "wave_c": 891.46, "tunnel_h": 912.43, "tunnel_l": 914.60, "ao": -30.49, "ao_sig": -30.10, "b76": 5.4, "b34": 7.2, "b16": 8.6, "b8": 14.4, "wr": 10.7, "rsi2": 12.9},
    },
    {
        "symbol": "MCX:MCXALUMDEX", "name": "MCX Aluminium Index", "last": 16274.86,
        "tf15": {"filter": 16265.13, "wave_h": 16265.18, "wave_l": 16232.04, "wave_c": 16249.59, "tunnel_h": 16178.89, "tunnel_l": 16150.90, "ao": 41.20, "ao_sig": 39.35, "b76": 67.8, "b34": 76.6, "b16": 51.9, "b8": 66.4, "wr": 50.0, "rsi2": 69.1},
        "tf1h": {"filter": 16255.16, "wave_h": 16237.46, "wave_l": 16143.87, "wave_c": 16201.43, "tunnel_h": 15846.66, "tunnel_l": 15800.95, "ao": 58.64, "ao_sig": 21.28, "b76": 80.2, "b34": 91.1, "b16": 64.7, "b8": 73.4, "wr": 73.9, "rsi2": 71.0},
        "tf4h": {"filter": 16163.41, "wave_h": 15973.89, "wave_l": 15764.45, "wave_c": 15887.55, "tunnel_h": 15330.01, "tunnel_l": 15234.50, "ao": 551.79, "ao_sig": 554.50, "b76": 82.8, "b34": 79.8, "b16": 78.4, "b8": 86.1, "wr": 68.6, "rsi2": 65.7},
    },
    {
        "symbol": "MCX:MCXMETLDEX", "name": "MCX Base Metal Index", "last": 21240.20,
        "tf15": {"filter": 21249.98, "wave_h": 21243.64, "wave_l": 21211.81, "wave_c": 21227.32, "tunnel_h": 21164.87, "tunnel_l": 21142.36, "ao": 43.47, "ao_sig": 50.91, "b76": 46.3, "b34": 58.8, "b16": 30.9, "b8": 25.3, "wr": 16.4, "rsi2": 11.0},
        "tf1h": {"filter": 21223.42, "wave_h": 21218.08, "wave_l": 21141.94, "wave_c": 21181.29, "tunnel_h": 20954.30, "tunnel_l": 20951.10, "ao": 50.85, "ao_sig": 13.12, "b76": 79.4, "b34": 63.9, "b16": 44.8, "b8": 57.2, "wr": 50.6, "rsi2": 45.2},
        "tf4h": {"filter": 21163.45, "wave_h": 21053.01, "wave_l": 20881.26, "wave_c": 20978.16, "tunnel_h": 21105.15, "tunnel_l": 21110.61, "ao": 486.40, "ao_sig": 531.34, "b76": 75.3, "b34": 85.6, "b16": 73.6, "b8": 59.7, "wr": 44.8, "rsi2": 52.8},
    },
    {
        "symbol": "MCX:MCXCOMPDEX", "name": "MCX Composite Index", "last": 24247.26,
        "tf15": {"filter": 24232.75, "wave_h": 24186.27, "wave_l": 24126.79, "wave_c": 24158.18, "tunnel_h": 24065.99, "tunnel_l": 24032.75, "ao": 146.59, "ao_sig": 166.18, "b76": 66.9, "b34": 88.6, "b16": 80.1, "b8": 43.1, "wr": 25.2, "rsi2": 34.9},
        "tf1h": {"filter": 24155.83, "wave_h": 24153.65, "wave_l": 24025.56, "wave_c": 24092.86, "tunnel_h": 23897.39, "tunnel_l": 23936.93, "ao": 35.16, "ao_sig": -53.55, "b76": 84.2, "b34": 66.7, "b16": 68.5, "b8": 88.3, "wr": 86.8, "rsi2": 95.8},
        "tf4h": {"filter": 24074.48, "wave_h": 24040.17, "wave_l": 23725.39, "wave_c": 23900.74, "tunnel_h": 24252.95, "tunnel_l": 24205.91, "ao": 709.42, "ao_sig": 802.02, "b76": 54.3, "b34": 87.3, "b16": 77.0, "b8": 55.8, "wr": 66.8, "rsi2": 71.7},
    },
    {
        "symbol": "MCX:MCXBULLDEX1!", "name": "MCX Bullion Futures", "last": 36044,
        "tf15": {"filter": 36270, "wave_h": 36098, "wave_l": 36076, "wave_c": 36093, "tunnel_h": 37354, "tunnel_l": 37576, "ao": 289, "ao_sig": 772, "b76": 33.7, "b34": 58.1, "b16": 12.9, "b8": 12.9, "wr": 29.9, "rsi2": 59.7},
        "tf1h": {"filter": 36183, "wave_h": 36195, "wave_l": 36146, "wave_c": 36182, "tunnel_h": 37823, "tunnel_l": 38036, "ao": 517, "ao_sig": 975, "b76": 31.0, "b34": 62.5, "b16": 16.0, "b8": 12.9, "wr": 28.4, "rsi2": 60.1},
        "tf4h": {"filter": 36046, "wave_h": 36790, "wave_l": 36619, "wave_c": 36728, "tunnel_h": 38124, "tunnel_l": 37996, "ao": -404, "ao_sig": -742, "b76": 32.8, "b34": 37.5, "b16": 73.4, "b8": 48.8, "wr": 28.4, "rsi2": 42.4},
    },
    {
        "symbol": "MCX:MCXMETLDEX1!", "name": "MCX Base Metal Futures", "last": 19000,
        "tf15": {"filter": None, "wave_h": None, "wave_l": None, "wave_c": None, "tunnel_h": None, "tunnel_l": None, "ao": None, "ao_sig": None, "b76": None, "b34": None, "b16": None, "b8": None, "wr": None, "rsi2": None},
        "tf1h": {"filter": None, "wave_h": None, "wave_l": None, "wave_c": None, "tunnel_h": None, "tunnel_l": None, "ao": None, "ao_sig": None, "b76": None, "b34": None, "b16": None, "b8": None, "wr": None, "rsi2": None},
        "tf4h": {"filter": None, "wave_h": None, "wave_l": None, "wave_c": None, "tunnel_h": None, "tunnel_l": None, "ao": None, "ao_sig": None, "b76": None, "b34": None, "b16": None, "b8": None, "wr": None, "rsi2": None},
    },
    {
        "symbol": "MCX:MCXBULLDEX", "name": "MCX Bullion Index", "last": 35676.49,
        "tf15": {"filter": 35684.71, "wave_h": 35678.79, "wave_l": 35545.48, "wave_c": 35614.52, "tunnel_h": 35732.25, "tunnel_l": 35707.67, "ao": 223.71, "ao_sig": 240.66, "b76": 38.3, "b34": 72.0, "b16": 59.8, "b8": 49.9, "wr": 19.8, "rsi2": 7.1},
        "tf1h": {"filter": 35645.18, "wave_h": 35900.83, "wave_l": 35615.84, "wave_c": 35749.21, "tunnel_h": 35814.77, "tunnel_l": 35963.14, "ao": -364.46, "ao_sig": -532.66, "b76": 59.4, "b34": 37.3, "b16": 38.1, "b8": 70.8, "wr": 69.7, "rsi2": 49.2},
        "tf4h": {"filter": 35777.96, "wave_h": 36056.44, "wave_l": 35422.77, "wave_c": 35772.57, "tunnel_h": 37375.57, "tunnel_l": 37414.55, "ao": 806.09, "ao_sig": 1150.85, "b76": 45.1, "b34": 75.4, "b16": 50.0, "b8": 33.0, "wr": 36.8, "rsi2": 44.6},
    },
    {
        "symbol": "MCX:MCXSILVDEX", "name": "MCX Silver Index", "last": 27451.77,
        "tf15": {"filter": 27388.82, "wave_h": 27341.41, "wave_l": 27196.35, "wave_c": 27271.25, "tunnel_h": 27497.33, "tunnel_l": 27495.08, "ao": 343.24, "ao_sig": 334.90, "b76": 42.6, "b34": 81.9, "b16": 72.9, "b8": 60.7, "wr": 35.0, "rsi2": 41.9},
        "tf1h": {"filter": 27310.88, "wave_h": 27669.71, "wave_l": 27352.48, "wave_c": 27503.44, "tunnel_h": 27729.54, "tunnel_l": 27889.40, "ao": -519.11, "ao_sig": -776.81, "b76": 53.4, "b34": 40.7, "b16": 40.7, "b8": 79.1, "wr": 81.9, "rsi2": 88.9},
        "tf4h": {"filter": 27598.25, "wave_h": 28026.88, "wave_l": 27305.03, "wave_c": 27701.80, "tunnel_h": 29718.81, "tunnel_l": 29845.70, "ao": 504.14, "ao_sig": 952.77, "b76": 38.7, "b34": 70.8, "b16": 43.2, "b8": 33.6, "wr": 42.1, "rsi2": 57.6},
    },
    {
        "symbol": "MCX:MCXCRUDEX", "name": "MCX Crude Oil Index", "last": 14314.28,
        "tf15": {"filter": 14248.45, "wave_h": 14176.42, "wave_l": 14027.24, "wave_c": 14111.64, "tunnel_h": 13525.00, "tunnel_l": 13445.59, "ao": 100.07, "ao_sig": 141.98, "b76": 80.8, "b34": 64.0, "b16": 75.0, "b8": 61.6, "wr": 83.3, "rsi2": 83.9},
        "tf1h": {"filter": 14045.15, "wave_h": 13687.36, "wave_l": 13446.29, "wave_c": 13593.41, "tunnel_h": 12737.24, "tunnel_l": 12598.28, "ao": 819.47, "ao_sig": 840.68, "b76": 88.0, "b34": 84.9, "b16": 83.0, "b8": 65.7, "wr": 67.3, "rsi2": 65.3},
        "tf4h": {"filter": 13459.43, "wave_h": 13037.56, "wave_l": 12578.61, "wave_c": 12826.25, "tunnel_h": 10886.19, "tunnel_l": 10589.45, "ao": 1111.40, "ao_sig": 734.61, "b76": 90.4, "b34": 91.9, "b16": 87.1, "b8": 87.0, "wr": 84.0, "rsi2": 96.7},
    },
    {
        "symbol": "MCX:MCXGOLDEX", "name": "MCX Gold Index", "last": 37626.74,
        "tf15": {"filter": 37659.42, "wave_h": 37672.28, "wave_l": 37543.51, "wave_c": 37610.79, "tunnel_h": 37679.33, "tunnel_l": 37646.60, "ao": 162.38, "ao_sig": 190.63, "b76": 36.3, "b34": 66.1, "b16": 52.7, "b8": 36.8, "wr": 16.9, "rsi2": 4.0},
        "tf1h": {"filter": 37637.27, "wave_h": 37841.13, "wave_l": 37573.36, "wave_c": 37700.05, "tunnel_h": 37697.62, "tunnel_l": 37836.05, "ao": -283.90, "ao_sig": -409.71, "b76": 62.0, "b34": 35.8, "b16": 36.8, "b8": 65.9, "wr": 62.8, "rsi2": 37.0},
        "tf4h": {"filter": 37701.46, "wave_h": 37923.13, "wave_l": 37337.27, "wave_c": 37651.39, "tunnel_h": 39108.32, "tunnel_l": 39153.62, "ao": 899.50, "ao_sig": 1183.98, "b76": 48.1, "b34": 77.3, "b16": 52.9, "b8": 32.9, "wr": 34.5, "rsi2": 40.3},
    },
]


def analyze_trend(d):
    if d["filter"] is None:
        return "N/A", "None", 0
    f = d["filter"]
    wc = d["wave_c"]
    th = d["tunnel_h"]
    tl = d["tunnel_l"]
    t_mid = (th + tl) / 2
    ao = d["ao"]
    b76 = d["b76"]

    # Determine alignment
    if f > wc and wc > t_mid:
        trend = "Bullish"
    elif f < wc and wc < t_mid:
        trend = "Bearish"
    elif f > wc and wc < t_mid:
        trend = "Recovering"
    elif f < wc and wc > t_mid:
        trend = "Weakening"
    else:
        trend = "Neutral"

    # Determine setup
    setup = "None"
    score = 3.0  # base

    if trend == "Bullish":
        # Check how far above tunnel
        if f > th:
            if b76 > 60 and d["b34"] > 60:
                setup = "PW Bull"
                score = 5.0
                if d["b16"] > 60 and d["b8"] > 60:
                    setup = "BO-2 Bull"
                    score = 6.0
            else:
                setup = "BO-1 Bull"
                score = 5.5
        else:
            setup = "FG Bull"
            score = 4.5
    elif trend == "Bearish":
        if f < tl:
            if b76 < 40 and d["b34"] < 40:
                setup = "PW Bear"
                score = 5.0
                if d["b16"] < 40 and d["b8"] < 40:
                    setup = "BO-2 Bear"
                    score = 6.0
            else:
                setup = "BO-1 Bear"
                score = 5.5
        else:
            setup = "FG Bear"
            score = 4.5
    elif trend == "Recovering":
        setup = "Recovery"
        score = 3.5
    elif trend == "Weakening":
        setup = "Pullback"
        score = 3.5

    # AO momentum modifier
    if ao is not None:
        if trend == "Bullish" and ao > 0:
            score += 1.0
        elif trend == "Bearish" and ao < 0:
            score += 1.0
        elif trend == "Bullish" and ao < 0:
            score -= 0.5
        elif trend == "Bearish" and ao > 0:
            score -= 0.5

    # Bungee stretch modifier
    if b76 is not None:
        if trend == "Bullish" and b76 > 70:
            score += 0.5
        elif trend == "Bearish" and b76 < 30:
            score += 0.5
        # Extreme stretch warning
        if b76 > 90 or b76 < 10:
            score += 0.5  # strong trend but risky

    score = max(0, min(10, score))
    return trend, setup, round(score, 1)


def mtf_score(d15, d1h, d4h):
    t15, s15, sc15 = analyze_trend(d15)
    t1h, s1h, sc1h = analyze_trend(d1h)
    t4h, s4h, sc4h = analyze_trend(d4h)

    # 4H has highest weight
    final = sc4h * 0.5 + sc1h * 0.3 + sc15 * 0.2

    # Confluence bonus
    directions = [t15, t1h, t4h]
    bull_count = sum(1 for d in directions if d in ["Bullish", "Recovering"])
    bear_count = sum(1 for d in directions if d in ["Bearish", "Weakening"])

    if bull_count == 3 or bear_count == 3:
        final += 1.0
    elif bull_count >= 2 or bear_count >= 2:
        final += 0.5
    # Opposing signals penalty
    if bull_count > 0 and bear_count > 0:
        final -= 0.5

    final = max(0, min(10, round(final, 1)))
    return t15, s15, sc15, t1h, s1h, sc1h, t4h, s4h, sc4h, final


def confidence_label(score):
    if score >= 8: return "HIGH"
    elif score >= 6: return "MED-HIGH"
    elif score >= 4: return "MEDIUM"
    elif score >= 2: return "LOW"
    else: return "PASS"


def score_fill(score):
    if score >= 7: return GREEN_FILL
    elif score >= 5: return YELLOW_FILL
    else: return RED_FILL


# ============================================================
# Sheet 1: Master Ranking
# ============================================================
ws1 = wb.active
ws1.title = "Master Ranking"
ws1.sheet_properties.tabColor = "BF8F00"

headers = ["Rank", "Symbol", "Name", "Last Price",
           "15m Trend", "15m Setup", "15m Score",
           "1H Trend", "1H Setup", "1H Score",
           "4H Trend", "4H Setup", "4H Score",
           "MTF Score", "Confidence", "Notes"]
ws1.append(headers)

results = []
for item in data:
    if item["tf4h"]["filter"] is None:
        results.append((item["symbol"], item["name"], item["last"],
                        "N/A", "None", 0, "N/A", "None", 0, "N/A", "None", 0, 0, "PASS", "Illiquid/No Data"))
        continue
    t15, s15, sc15, t1h, s1h, sc1h, t4h, s4h, sc4h, mtf = mtf_score(item["tf15"], item["tf1h"], item["tf4h"])
    note = ""
    if sc4h >= 7 and t4h == "Bullish":
        note = "Strong bullish trend"
    elif sc4h >= 7 and t4h == "Bearish":
        note = "Strong bearish trend"
    elif mtf >= 7:
        note = "Multi-TF confluence"
    b76_4h = item["tf4h"]["b76"]
    if b76_4h and (b76_4h > 85 or b76_4h < 15):
        note += " | STRETCHED" if note else "STRETCHED"
    if t4h != t1h and t4h != "Neutral":
        note += " | TF conflict" if note else "TF conflict"
    results.append((item["symbol"], item["name"], item["last"],
                    t15, s15, sc15, t1h, s1h, sc1h, t4h, s4h, sc4h, mtf,
                    confidence_label(mtf), note))

# Sort by MTF score descending
results.sort(key=lambda x: x[12] if isinstance(x[12], (int, float)) else 0, reverse=True)

for rank, r in enumerate(results, 1):
    ws1.append([rank] + list(r))

style_headers(ws1, len(headers), DARK_GOLD)
style_body(ws1, len(results) + 1, len(headers))

# Color code scores
for row in range(2, len(results) + 2):
    for col in [7, 10, 13, 14]:  # 15m/1H/4H/MTF score columns
        cell = ws1.cell(row=row, column=col)
        try:
            val = float(cell.value)
            cell.fill = score_fill(val)
        except (ValueError, TypeError):
            pass
    # Confidence column
    conf_cell = ws1.cell(row=row, column=15)
    if conf_cell.value == "HIGH":
        conf_cell.fill = GREEN_FILL
    elif conf_cell.value == "MED-HIGH":
        conf_cell.fill = PatternFill('solid', fgColor='B4C6E7')
    elif conf_cell.value == "MEDIUM":
        conf_cell.fill = YELLOW_FILL
    elif conf_cell.value in ["LOW", "PASS"]:
        conf_cell.fill = RED_FILL

auto_width(ws1, len(headers))
ws1.freeze_panes = "A2"


# ============================================================
# Sheet 2: 15m Technical Detail
# ============================================================
ws2 = wb.create_sheet("15m Detail")
ws2.sheet_properties.tabColor = "548235"

detail_headers = ["Symbol", "Name", "Last", "Filter(12)", "Wave H(34)", "Wave L(34)", "Wave C(34)",
                   "Tunnel(144)", "Tunnel(169)", "AO", "AO Sig", "Bungee 76", "Bungee 34",
                   "Bungee 16", "Bungee 8", "Williams%R", "RSI(2)", "Trend", "Setup", "Score"]
ws2.append(detail_headers)

for item in data:
    d = item["tf15"]
    if d["filter"] is None:
        ws2.append([item["symbol"], item["name"], item["last"]] + ["N/A"] * 17)
        continue
    t, s, sc = analyze_trend(d)
    ws2.append([item["symbol"], item["name"], item["last"],
                d["filter"], d["wave_h"], d["wave_l"], d["wave_c"],
                d["tunnel_h"], d["tunnel_l"], d["ao"], d["ao_sig"],
                d["b76"], d["b34"], d["b16"], d["b8"], d["wr"], d["rsi2"],
                t, s, sc])

style_headers(ws2, len(detail_headers))
style_body(ws2, len(data) + 1, len(detail_headers))
auto_width(ws2, len(detail_headers))
ws2.freeze_panes = "C2"


# ============================================================
# Sheet 3: 1H Technical Detail
# ============================================================
ws3 = wb.create_sheet("1H Detail")
ws3.sheet_properties.tabColor = "2E75B6"

ws3.append(detail_headers)

for item in data:
    d = item["tf1h"]
    if d["filter"] is None:
        ws3.append([item["symbol"], item["name"], item["last"]] + ["N/A"] * 17)
        continue
    t, s, sc = analyze_trend(d)
    ws3.append([item["symbol"], item["name"], item["last"],
                d["filter"], d["wave_h"], d["wave_l"], d["wave_c"],
                d["tunnel_h"], d["tunnel_l"], d["ao"], d["ao_sig"],
                d["b76"], d["b34"], d["b16"], d["b8"], d["wr"], d["rsi2"],
                t, s, sc])

style_headers(ws3, len(detail_headers))
style_body(ws3, len(data) + 1, len(detail_headers))
auto_width(ws3, len(detail_headers))
ws3.freeze_panes = "C2"


# ============================================================
# Sheet 4: 4H Technical Detail
# ============================================================
ws4 = wb.create_sheet("4H Detail")
ws4.sheet_properties.tabColor = "BF8F00"

ws4.append(detail_headers)

for item in data:
    d = item["tf4h"]
    if d["filter"] is None:
        ws4.append([item["symbol"], item["name"], item["last"]] + ["N/A"] * 17)
        continue
    t, s, sc = analyze_trend(d)
    ws4.append([item["symbol"], item["name"], item["last"],
                d["filter"], d["wave_h"], d["wave_l"], d["wave_c"],
                d["tunnel_h"], d["tunnel_l"], d["ao"], d["ao_sig"],
                d["b76"], d["b34"], d["b16"], d["b8"], d["wr"], d["rsi2"],
                t, s, sc])

style_headers(ws4, len(detail_headers))
style_body(ws4, len(data) + 1, len(detail_headers))

# Color code score column (col 20) and trend column (col 18)
for row in range(2, len(data) + 2):
    sc_cell = ws4.cell(row=row, column=20)
    try:
        val = float(sc_cell.value)
        sc_cell.fill = score_fill(val)
    except (ValueError, TypeError):
        pass
    trend_cell = ws4.cell(row=row, column=18)
    if trend_cell.value == "Bullish":
        trend_cell.fill = GREEN_FILL
    elif trend_cell.value == "Bearish":
        trend_cell.fill = RED_FILL
    elif trend_cell.value in ["Recovering", "Weakening"]:
        trend_cell.fill = YELLOW_FILL

auto_width(ws4, len(detail_headers))
ws4.freeze_panes = "C2"


# ============================================================
# Sheet 5: Setup Legend & Confidence Scale
# ============================================================
ws5 = wb.create_sheet("Legend")
ws5.sheet_properties.tabColor = "7030A0"

ws5.append(["Setup", "Type", "Description", "Entry Condition"])
legend_data = [
    ["PW", "Trend Following", "Price Walk - price rides along the tunnel", "Filter > Wave > Tunnel (bull) or reverse (bear), Bungee aligned"],
    ["FG", "Counter-Trend", "Fading the Ghost - retracement after PW", "Wave 2/B correction, reversal candle + Filter close confirms"],
    ["BO-1", "Trend Following", "Breakout 1 - Wave crosses Tunnel", "Wave crosses Tunnel, Filter/Wave/Tunnel converge, candle closes beyond Filter"],
    ["BO-2", "Trend Following", "Breakout 2 - Wave bounces off Tunnel", "Wave bounces off Tunnel after correction, re-entry into existing trend"],
    ["BO-3", "Trend Following", "Breakout 3 - Pullback to Wave", "Price pulls back to Wave (34 EMA), AO pullback towards zero, Bungee SnapBack"],
    ["BO-4", "Trend Following", "Breakout 4 - Pullback to Tunnel", "Deepest pullback to Tunnel itself, AO near/at zero, all Bungee strands snap back"],
]
for row in legend_data:
    ws5.append(row)

ws5.append([])
ws5.append(["Confidence Scale"])
ws5.append(["Score Range", "Label", "Color", "Action"])
scale_data = [
    ["8.0 - 10.0", "HIGH", "Green", "Active trade setup - execute per SOP"],
    ["6.0 - 7.9", "MED-HIGH", "Blue", "Good setup forming - prepare entry"],
    ["4.0 - 5.9", "MEDIUM", "Yellow", "Watch list - monitor for development"],
    ["2.0 - 3.9", "LOW", "Red", "No clear setup - avoid"],
    ["0.0 - 1.9", "PASS", "Red", "Opposing trend / no data"],
]
for row in scale_data:
    ws5.append(row)

ws5.append([])
ws5.append(["Indicator Mapping"])
ws5.append(["Study Output", "Wavy Tunnel Component", "Period"])
mapping_data = [
    ["Fast EMA (12)", "Filter", "12"],
    ["Tunnel High/Low/Close (34)", "Wave", "34 (HL/2)"],
    ["Wave EMA (144/169)", "Tunnel", "144 / 169"],
    ["Awesome Oscillator", "AO", "5/34 SMA diff"],
    ["Slower Stochastic (76)", "Bungee Red (slowest)", "76"],
    ["Slow Stochastic (34)", "Bungee Orange", "34"],
    ["Medium Stochastic (16)", "Bungee Green", "16"],
    ["Fast Stochastic (8)", "Bungee Blue (fastest)", "8"],
]
for row in mapping_data:
    ws5.append(row)

style_headers(ws5, 4)
auto_width(ws5, 4)
# Bold sub-headers
for row in ws5.iter_rows():
    for cell in row:
        if cell.value in ["Confidence Scale", "Indicator Mapping"]:
            cell.font = Font(name='Arial', bold=True, size=12, color='1F4E79')

output = "C:\\Users\\91700\\tradingview-mcp-jackson\\Gold_Watchlist_Analysis_2026-04-04.xlsx"
wb.save(output)
print(f"Saved to {output}")
